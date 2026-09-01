import pandas as pd
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Connect to database
conn = sqlite3.connect('nasa_cneos.db')

# Extract full dataset
df = pd.read_sql_query("SELECT * FROM fact_asteroid_approaches", conn)
conn.close()

# Create Excel writer
file_path = "Asteroid_Executive_Risk_Report.xlsx"
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Data_Cleansed', index=False)

# Load workbook for styling
wb = openpyxl.load_workbook(file_path)
ws_data = wb['Data_Cleansed']

# Style Data Sheet
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col_num, col_name in enumerate(df.columns, 1):
    cell = ws_data.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Auto-fit columns
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(file_path)
print(f"Excel report successfully generated: {file_path}")